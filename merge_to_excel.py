import pandas as pd
import importlib.util
import sys
import os
import shutil

def convert_pdf_to_csv(pdf_path, converter_script):
    """
    Helper function to convert PDF to CSV using a dynamic converter script.
    Returns the path to the generated CSV file.
    """
    if not os.path.exists(converter_script):
        raise FileNotFoundError(f"Converter script {converter_script} not found.")

    # Determine proper output path in output/ folder
    # e.g. convert_globe.py -> output/Globe/filename.csv
    script_basename = os.path.basename(converter_script)
    name_part = script_basename.replace(".py", "")
    if name_part.startswith("convert_"):
        name_part = name_part.replace("convert_", "")
    
    subfolder = name_part.capitalize()
    
    output_dir = os.path.join("output", subfolder)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    filename = os.path.basename(pdf_path).replace(".pdf", ".csv")
    csv_path = os.path.join(output_dir, filename)

    # Load converter module dynamically
    spec = importlib.util.spec_from_file_location("dynamic_converter", converter_script)
    module = importlib.util.module_from_spec(spec)
    sys.modules["dynamic_converter"] = module
    spec.loader.exec_module(module)

    if not hasattr(module, 'process_file'):
        raise AttributeError("Converter script must have a 'process_file' function.")

    print(f"Converting {pdf_path} using {converter_script}...")
    success = module.process_file(pdf_path, csv_path)
    
    if not success:
        # Check if file exists anyway
        if os.path.exists(csv_path) and os.path.getsize(csv_path) > 0:
            print("Warning: Converter returned False but CSV exists. Continuing...")
        else:
            raise Exception("PDF to CSV conversion failed.")
            
    return csv_path

def extract_data_from_csv(csv_path, pdf_cols):
    """
    Reads the CSV and extracts the first column (assumed Key) and the specified pdf_cols.
    Returns a DataFrame with the unique key column and requested data columns.
    """
    print(f"Reading converted data from {csv_path}...")
    # Read CSV (Force all to string to preserve formats like '0004')
    df = pd.read_csv(csv_path, dtype=str)
    
    if df.empty:
        raise Exception("The converted CSV file is empty.")

    # The first column is assumed to be the Key (Item Number)
    key_col = df.columns[0]
    print(f"Using first column '{key_col}' as the Key.")

    # Check which requested columns are actually present and handle simple aliases if needed
    # (For now, we just filter for existing columns to avoid errors)
    present_cols = []
    missing_cols = []
    
    for col in pdf_cols:
        if col in df.columns:
            present_cols.append(col)
        else:
            missing_cols.append(col)
            
    if missing_cols:
        print(f"WARNING: The following requested columns were not found in the CSV: {missing_cols}")
        print(f"Available columns: {list(df.columns)}")

    # Construct final list of columns to extract: [Key] + [Found Requested Cols]
    # Ensure no duplicates if the Key was also requested in pdf_cols
    final_cols = [key_col]
    for col in present_cols:
        if col != key_col:
            final_cols.append(col)
            
    return df[final_cols]

def find_element_row_in_excel(excel_path, element):
    """
    Scans the Excel file to find the row index of a specific element.
    Returns the 0-based row index, or -1 if not found.
    """
    # print(f"Scanning Excel {excel_path} for row of element: {element}...")
    try:
        # Scan whole file (removed nrows limit)
        df_scan = pd.read_excel(excel_path, header=None, dtype=str)
    except Exception as e:
        raise Exception(f"Failed to read Excel file: {e}")
    
    for r in range(len(df_scan)):
        for c in range(len(df_scan.columns)):
            cell_val = df_scan.iloc[r, c]
            if pd.isna(cell_val): continue
            if str(cell_val).strip() == element:
                return r
    return -1

def find_element_col_in_excel(excel_path, element):
    """
    Scans the Excel file to find the column index of a specific element.
    Returns the 0-based column index, or -1 if not found.
    """
    # print(f"Scanning Excel {excel_path} for col of element: {element}...")
    try:
        # Scan whole file (removed nrows limit)
        df_scan = pd.read_excel(excel_path, header=None, dtype=str)
    except Exception as e:
        raise Exception(f"Failed to read Excel file: {e}")
    
    for r in range(len(df_scan)):
        for c in range(len(df_scan.columns)):
            cell_val = df_scan.iloc[r, c]
            if pd.isna(cell_val): continue
            if str(cell_val).strip() == element:
                return c
    return -1

def process_merge(pdf_path, converter_script, excel_path, pdf_cols, excel_cols, output_path=None):
    print("=== Starting Merge Process ===")
    print(f"PDF Path: {pdf_path}")
    print(f"Converter Script: {converter_script}")
    print(f"Excel Path: {excel_path}")
    print(f"Mapping: {pdf_cols} -> {excel_cols}")

    # 1. Convert PDF to CSV and get path
    try:
        print("\n--- Step 1: Converting PDF to CSV ---")
        csv_path = convert_pdf_to_csv(pdf_path, converter_script)
        print(f"Successfully converted PDF to CSV: {csv_path}")
    except Exception as e:
        print(f"FAILED at Step 1 (PDF Conversion): {e}")
        raise

    # 2. Extract Data from CSV
    # Returns DataFrame with [Key, Col1, Col2...]
    try:
        print("\n--- Step 2: Extracting Data from CSV ---")
        df_data = extract_data_from_csv(csv_path, pdf_cols)
        print(f"Extracted {len(df_data)} rows from CSV.")
        if not df_data.empty:
            print(f"Sample keys: {df_data.iloc[:3, 0].tolist()}")
    except Exception as e:
        print(f"FAILED at Step 2 (Data Extraction): {e}")
        raise
    
    # 3. Prepare to update Excel
    # We use openpyxl to update the file while preserving formatting
    ws = None
    wb = None
    try:
        print(f"\n--- Step 3: Loading Excel File: {excel_path} ---")
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        print("Excel file loaded successfully.")
    except ImportError:
        print("FAILED at Step 3: openpyxl not found.")
        raise ImportError("openpyxl is required. Please install it with 'pip install openpyxl'")
    except Exception as e:
        print(f"FAILED at Step 3 (Loading Excel): {e}")
        raise

    # Pre-find the column indices for the target excel columns
    # We do this once upfront
    print("\n--- Step 4: Locating Target Columns in Excel ---")
    excel_col_indices = {}
    for col_name in excel_cols:
        print(f"Looking for column '{col_name}'...")
        col_idx = find_element_col_in_excel(excel_path, col_name)
        if (col_idx != -1):
            excel_col_indices[col_name] = col_idx
            print(f"  -> Found '{col_name}' at column index {col_idx}")
        else:
            print(f"  -> WARNING: Target column '{col_name}' NOT found in Excel.")

    if not excel_col_indices:
        print("WARNING: No target columns were found in the Excel file. No updates will be possible.")

    # 4. Iterate iterate through items and update
    updates_count = 0
    key_col_name = df_data.columns[0] # The first column is the key (Item Number)
    
    print("\n--- Step 5: Mapping Data and Updating Rows ---")
    total_items = len(df_data)
    processed_items = 0
    
    for _, row in df_data.iterrows():
        processed_items += 1
        item_val = str(row[key_col_name]).strip()
        
        # Log progress every 10 items or for the first few
        if processed_items <= 5 or processed_items % 10 == 0:
             print(f"Processing Item {processed_items}/{total_items}: '{item_val}'...")

        # Find which row this item is on in the Excel file
        try:
            row_idx = find_element_row_in_excel(excel_path, item_val)
        except Exception as e:
            print(f"Error searching for item '{item_val}': {e}")
            continue
        
        if row_idx != -1:
            # We found the item, now update the specific columns
            # pdf_cols maps to excel_cols by index
            print(f"  -> Match found at Row index {row_idx}. Updating...")
            for i in range(len(pdf_cols)):
                p_col = pdf_cols[i]
                e_col = excel_cols[i]
                
                # If we know where the target column is
                if e_col in excel_col_indices:
                    val_to_write = row[p_col]
                    
                    # Convert 0-based index to 1-based OpenPyXL index
                    target_row = row_idx + 1
                    target_col = excel_col_indices[e_col] + 1
                    
                    # Log cleanup
                    # print(f"Updating {item_val} at ({target_row}, {target_col}): {val_to_write}")
                    
                    # Number Logic: Handle European vs US format
                    # The CSV likely contains "1.620,00" (EU) or "1,620.00" (US)
                    # We attempt to parse this intelligently.
                    try:
                        if isinstance(val_to_write, str):
                            clean_val = val_to_write.strip()
                            # Only attempt conversion if it looks like a number
                            # (Starts with digit, -, or +)
                            if len(clean_val) > 0 and (clean_val[0].isdigit() or clean_val[0] in '-+'):
                                if ',' in clean_val and '.' in clean_val:
                                    last_dot = clean_val.rfind('.')
                                    last_comma = clean_val.rfind(',')
                                    
                                    if last_comma > last_dot:
                                        # EU Format: 1.000,00 (Comma is decimal)
                                        clean_val = clean_val.replace('.', '').replace(',', '.')
                                    else:
                                        # US Format: 1,000.00 (Dot is decimal)
                                        clean_val = clean_val.replace(',', '')
                                elif ',' in clean_val:
                                     # Ambiguous case: "1,000" or "0,524"
                                     # Given the previous context of "1.620,00", this file uses EU format.
                                     # So we treat ',' as decimal.
                                     clean_val = clean_val.replace(',', '.')
                                     
                                val_to_write = float(clean_val)
                    except Exception as e:
                        # Conversion failed, keeping original string value
                        pass

                    try:
                        ws.cell(row=target_row, column=target_col, value=val_to_write)
                    except Exception as e:
                        print(f"Error writing to cell ({target_row}, {target_col}): {e}")
            
            updates_count += 1
        else:
            if processed_items <= 5: # Only log failures for the first few to avoid spam
                 print(f"  -> Item '{item_val}' NOT found in Excel.")
            pass

    # 5. Save the result
    print("\n--- Step 6: Saving Output ---")
    if not output_path:
        output_path = excel_path.replace(".xlsx", "_merged.xlsx")
        
    try:
        wb.save(output_path)
        print(f"Done. Saved merged file to {output_path}")
        print(f"Summary: Updated {updates_count} out of {total_items} items.")
    except Exception as e:
        print(f"FAILED at Step 6 (Saving File): {e}")
        raise

    return output_path

if __name__ == "__main__":
    sample_pdf = "source/Globe/51075 Globe 124652.pdf"
    sample_converter = "convert_globe.py"
    sample_excel = "source/124652.xlsx"
    
    list_from_pdf = ["Quantity", "Price (USD)"] # Updated to match Detected Headers
    list_to_excel = ["数量", "单价"] 
    
    try:
        process_merge(sample_pdf, sample_converter, sample_excel, list_from_pdf, list_to_excel)
    except Exception as e:
        print(f"Error: {e}")
