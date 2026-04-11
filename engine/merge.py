"""
Insertion engine — merges extracted CSV data into an Excel template.

This module is **PDF-agnostic**: it receives a path to a CSV file (already
produced by a converter in the ``converters`` package) and writes selected
columns into an Excel workbook, matching rows by a key column.
"""

import logging
import pandas as pd
import os

logger = logging.getLogger(__name__)


class NoMatchesFoundError(Exception):
    """Raised when no CSV keys match any row in the Excel template."""


# ── Data extraction ───────────────────────────────────────────────────

def extract_data_from_csv(csv_path, data_cols):
    """
    Read the CSV and extract the key column (first column) plus *data_cols*.

    Returns a DataFrame with ``[key_col, *found_data_cols]``.
    """
    logger.info("Reading converted data from %s", csv_path)
    df = pd.read_csv(csv_path, dtype=str)

    if df.empty:
        raise Exception("The converted CSV file is empty.")

    key_col = df.columns[0]
    logger.info("Using first column '%s' as the Key.", key_col)

    present_cols = [c for c in data_cols if c in df.columns]
    missing_cols = [c for c in data_cols if c not in df.columns]

    if missing_cols:
        logger.warning("Columns not found in CSV: %s", missing_cols)
        logger.warning("Available columns: %s", list(df.columns))

    final_cols = [key_col] + [c for c in present_cols if c != key_col]
    return df[final_cols]


# ── Excel lookup (read once) ─────────────────────────────────────────

def _build_excel_lookup(excel_path):
    """
    Read the Excel file *once* and return two dicts:

    * ``row_lookup``: ``cell_value → first 0-based row index``
    * ``col_lookup``: ``cell_value → first 0-based column index``

    Duplicate values record only the **first** (top-left) occurrence.
    """
    df = pd.read_excel(excel_path, header=None, dtype=str)

    row_lookup = {}
    col_lookup = {}

    for r in range(len(df)):
        for c in range(len(df.columns)):
            cell = df.iloc[r, c]
            if pd.isna(cell):
                continue
            val = str(cell).strip()
            if val not in row_lookup:
                row_lookup[val] = r
            if val not in col_lookup:
                col_lookup[val] = c

    return row_lookup, col_lookup


# ── Number format helpers ─────────────────────────────────────────────

def _smart_number_convert(value):
    """
    Convert a string to float, handling EU (``1.000,50``) and US
    (``1,000.50``) number formats transparently.

    Returns the original value unchanged when conversion fails.
    """
    if not isinstance(value, str):
        return value

    clean = value.strip()
    if not clean or not (clean[0].isdigit() or clean[0] in "-+"):
        return value

    try:
        if "," in clean and "." in clean:
            if clean.rfind(",") > clean.rfind("."):
                # EU: 1.000,00
                clean = clean.replace(".", "").replace(",", ".")
            else:
                # US: 1,000.00
                clean = clean.replace(",", "")
        elif "," in clean:
            # Assume EU decimal comma
            clean = clean.replace(",", ".")
        return float(clean)
    except (ValueError, IndexError):
        return value


# ── Public API ────────────────────────────────────────────────────────

def process_merge(csv_path, excel_path, csv_cols, excel_cols, output_path=None):
    """
    Merge columns from a converted CSV into an Excel template.

    The **first column** of the CSV is used as the lookup key.  For every
    row the engine finds the key value in the Excel sheet, then writes the
    requested *csv_cols* values into the corresponding *excel_cols*.

    Args:
        csv_path:    Path to the source CSV (output of a converter).
        excel_path:  Path to the Excel template.
        csv_cols:    List of CSV column names to read.
        excel_cols:  List of Excel column headers to write into (same order).
        output_path: Where to save the result (default: ``*_merged.xlsx``).

    Returns:
        Path to the saved output file.
    """
    logger.info("=== Starting Merge Process ===")
    logger.info("CSV Path:   %s", csv_path)
    logger.info("Excel Path: %s", excel_path)
    logger.info("Mapping:    %s → %s", csv_cols, excel_cols)

    # ── 1. Extract data from CSV ──────────────────────────────────────
    try:
        logger.info("--- Step 1: Extracting Data from CSV ---")
        df_data = extract_data_from_csv(csv_path, csv_cols)
        logger.info("Extracted %d rows.", len(df_data))
        if not df_data.empty:
            logger.debug("Sample keys: %s", df_data.iloc[:3, 0].tolist())
    except Exception as e:
        logger.error("FAILED at Step 1 (Data Extraction): %s", e)
        raise

    # ── 2. Load Excel workbook ────────────────────────────────────────
    try:
        logger.info("--- Step 2: Loading Excel File ---")
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        logger.info("Excel file loaded successfully.")
    except ImportError:
        raise ImportError(
            "openpyxl is required. Install with: pip install openpyxl")
    except Exception as e:
        logger.error("FAILED at Step 2 (Loading Excel): %s", e)
        raise

    # ── 3. Build lookup maps (single Excel read) ──────────────────────
    logger.info("--- Step 3: Building Excel Lookup Maps ---")
    row_lookup, col_lookup = _build_excel_lookup(excel_path)

    excel_col_indices = {}
    for col_name in excel_cols:
        if col_name in col_lookup:
            excel_col_indices[col_name] = col_lookup[col_name]
            logger.debug("  Found column '%s' at index %d",
                         col_name, col_lookup[col_name])
        else:
            logger.warning("  Column '%s' NOT found in Excel.", col_name)

    if not excel_col_indices:
        logger.warning("No target columns found. No updates possible.")

    # ── 4. Map CSV rows → Excel rows and update cells ─────────────────
    key_col_name = df_data.columns[0]
    total = len(df_data)
    updates = 0

    logger.info("--- Step 4: Updating %d Items ---", total)
    for idx, (_, row) in enumerate(df_data.iterrows(), 1):
        item_val = str(row[key_col_name]).strip()

        if idx <= 5 or idx % 10 == 0:
            logger.debug("Processing %d/%d: '%s'", idx, total, item_val)

        if item_val not in row_lookup:
            if idx <= 5:
                logger.debug("  → Item '%s' NOT found in Excel.", item_val)
            continue

        row_idx = row_lookup[item_val]
        target_row = row_idx + 1  # openpyxl is 1-indexed

        for c_col, e_col in zip(csv_cols, excel_cols):
            if e_col not in excel_col_indices:
                continue
            target_col = excel_col_indices[e_col] + 1  # 1-indexed
            val = _smart_number_convert(row.get(c_col, ""))

            try:
                ws.cell(row=target_row, column=target_col, value=val)
            except Exception as e:
                logger.error("  Error writing (%d, %d): %s",
                             target_row, target_col, e)

        updates += 1

    if updates == 0:
        msg = "No matches found. No output file was generated."
        logger.warning(msg)
        raise NoMatchesFoundError(msg)

    # ── 5. Save ───────────────────────────────────────────────────────
    if not output_path:
        output_path = excel_path.replace(".xlsx", "_merged.xlsx")

    logger.info("--- Step 5: Saving to %s ---", output_path)
    try:
        wb.save(output_path)
        logger.info("Done. Updated %d/%d items.", updates, total)
    except Exception as e:
        logger.error("FAILED at Step 5 (Saving): %s", e)
        raise

    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    # Quick manual test
    import sys
    if len(sys.argv) >= 5:
        csv_p = sys.argv[1]
        excel_p = sys.argv[2]
        c_cols = [c.strip() for c in sys.argv[3].split(",")]
        e_cols = [c.strip() for c in sys.argv[4].split(",")]
        out = sys.argv[5] if len(sys.argv) > 5 else None
        process_merge(csv_p, excel_p, c_cols, e_cols, output_path=out)
    else:
        logger.info("Usage: python -m engine.merge <csv> <excel> <csv_cols> <excel_cols> [output]")
