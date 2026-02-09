import pandas as pd
from openpyxl import load_workbook
import sys

excel_path = '/Users/tim/Documents/dad/simpletranslator/source/124652.xlsx'

print(f"Reading {excel_path}...")

# 1. Replicate Pandas Header Detection
df_raw = pd.read_excel(excel_path, header=None, dtype=str)
# Mock CSV Items (from seeing PDF output)
csv_items = set(['000460150000000', '000460150000102', '000460150001802', '000460150003802'])

best_header_row = 0
best_overlap = 0
best_header_score = -1
best_key = ""

# excel_cols definition for simulation
excel_cols = ["数量", "单价", "合计"]

print("--- Header Detection ---")
for r in range(10):
    row_vals = df_raw.iloc[r]
    data = df_raw.iloc[r+1:]
    
    # 1. Data Overlap
    max_ov = 0
    best_c = ""
    best_c_idx = 0
    
    for c in range(len(row_vals)):
        col_vals = set(data.iloc[:, c].dropna())
        overlap = len(csv_items.intersection(col_vals))
        if overlap > max_ov:
            max_ov = overlap
            best_c = row_vals.iloc[c]
            best_c_idx = c
            
    # 2. Keyword Match
    current_header_match_count = 0
    row_values_str = [str(x).strip() for x in row_vals.values if pd.notna(x)]
    
    for target_col in excel_cols:
        if target_col in row_values_str:
            current_header_match_count += 1
            
    # 3. Explicit Key Search
    possible_key_names = ['Item', 'Part Number', '客人的新货号', '货号']
    found_key_name = None
    for val in row_values_str:
        if val in possible_key_names:
            found_key_name = val
            current_header_match_count += 2
            break
            
    # If key found by name, ensure we pick it up
    if found_key_name and max_ov == 0:
        best_c = found_key_name
        # Find index
        for idx, val in enumerate(row_vals):
             if str(val).strip() == found_key_name:
                 best_c_idx = idx
                 break
        max_ov = 1

    print(f"Row {r}: Matches={current_header_match_count}, Overlap={max_ov}, BestCol='{best_c}'")
    
    # Decision
    is_better = False
    if current_header_match_count > best_header_score:
        is_better = True
    elif current_header_match_count == best_header_score:
        if max_ov > best_overlap:
            is_better = True
        elif max_ov == best_overlap:
             if str(best_c).lower() != 'nan' and str(best_key) == 'nan':
                  is_better = True

    if is_better:
        best_overlap = max_ov
        best_header_row = r
        best_key = best_c
        best_header_score = current_header_match_count

print(f"Winner: Row {best_header_row}, Key '{best_key}'")

# 2. Replicate OpenPyXL Mapping
print("\n--- OpenPyXL Mapping ---")
wb = load_workbook(excel_path)
ws = wb.active
header_row_idx = best_header_row + 1

print(f"Reading row {header_row_idx} via openpyxl...")
col_map = {}
for cell in ws[header_row_idx]:
    if cell.value:
        val_str = str(cell.value).strip()
        col_map[val_str] = cell.column
        print(f"  Col {cell.column}: '{val_str}' (Repr: {repr(val_str)})")

target_col = "数量"
if target_col in col_map:
    print(f"Found Target '{target_col}' at Column {col_map[target_col]}")
else:
    print(f"FAILED TO FIND '{target_col}' in map keys: {list(col_map.keys())}")

# 3. Check Row Values
print("\n--- Checking Data Row ---")
# Find a row with Key = '000460150000102' (from earlier inspection)
key_col_idx = col_map.get(best_key, None)
if key_col_idx:
    for row in ws.iter_rows(min_row=header_row_idx+1):
        k_val = str(row[key_col_idx-1].value).strip() if row[key_col_idx-1].value else ""
        if k_val == '000460150000102':
            print(f"Found Row for '000460150000102'")
            target_idx = col_map.get(target_col)
            if target_idx:
                curr_val = row[target_idx-1].value
                print(f"Current Value in '{target_col}': {curr_val}")
            break
else:
    print(f"Cannot scan rows, missing key column '{best_key}' index")
