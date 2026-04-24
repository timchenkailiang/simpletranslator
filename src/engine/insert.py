"""
Insertion engine — inserts extracted CSV data into an Excel template.

This module is **PDF-agnostic**: it receives a path to a CSV file (already
produced by a converter in the ``converters`` package) and writes selected
columns into an Excel workbook, matching rows by a key column.
"""

import logging
import math
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from utils import smart_number_convert

logger = logging.getLogger(__name__)

# Excel number_format code that forces a cell to be treated as text.  We
# apply it to identifier-like values (leading-zero part numbers, long
# digit strings) so Excel does not strip the zeros or switch to
# scientific notation on display.
_TEXT_FORMAT = "@"

# Every cell written by the insert engine is stamped with this font so
# the merged sheet has a consistent typography regardless of whatever
# (possibly mixed) fonts the template happened to carry.
_INSERT_FONT_NAME = "Times New Roman"
_INSERT_FONT_SIZE = 9

# Red fill — applied to all cells that were updated (written to)
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                        fill_type="solid")
# Yellow fill — applied to cells flagged by validation (warnings)
_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00",
                           fill_type="solid")


class NoMatchesFoundError(Exception):
    """Raised when no CSV keys match any row in the Excel template."""
    def __init__(self, message, stats=None):
        super().__init__(message)
        self.stats = stats or {}


# ── Quantity recalculation ─────────────────────────────────────────────

def _normalize_tolerance(value, quantity):
    """Convert a user-provided tolerance to a **ratio** (float).

    Accepted formats:
      * ``"15%"`` or ``" 15 % "`` → 0.15  (percentage string)
      * ``0.15``                   → 0.15  (number < 1 treated as ratio)
      * ``100``                    → ratio = 100 / quantity  (exact piece count)

    Returns 0.0 when *quantity* ≤ 0 and an exact count is given.
    """
    if isinstance(value, str):
        value = value.strip()
        if value.endswith("%"):
            return float(value.rstrip("%").strip()) / 100.0
        value = float(value)

    value = float(value)
    if value < 1:
        # Already a ratio (e.g. 0.15 = 15 %)
        return value
    # Exact piece count — convert to ratio relative to quantity
    if quantity <= 0:
        return 0.0
    return value / quantity


def calculate_new_quantity(quantity, pcs_per_ctn, pcs_per_plt,
                           max_increase=0.15, max_decrease=0.15):
    """
    Adjust *quantity* to fit efficiently into pallets and cartons.

    The function may increase **or** decrease the quantity within a
    user-controlled range to maximise full-pallet usage and minimise
    loose cartons.

    Priority order:
      1. Full pallets (round up, then round down).
      2. Mixed packing — maximise pallets, then minimise cartons.

    Args:
        quantity:       Original quantity from the PDF/CSV.
        pcs_per_ctn:    Pieces per carton (from Excel template).
        pcs_per_plt:    Pieces per pallet (from Excel template).
        max_increase:   Maximum allowed increase.  Accepts:
                        ``0.15`` (ratio), ``"15%"`` (percentage string),
                        or ``100`` (exact piece count ≥ 1).
        max_decrease:   Maximum allowed decrease (same formats).

    Returns:
        The adjusted quantity (int).
    """
    quantity = int(quantity)
    pcs_per_ctn = int(pcs_per_ctn)
    pcs_per_plt = int(pcs_per_plt)

    # Guard against invalid packing info
    if pcs_per_ctn <= 0 or pcs_per_plt <= 0 or quantity <= 0:
        return quantity

    # Normalise tolerances to ratios
    increase_ratio = _normalize_tolerance(max_increase, quantity)
    decrease_ratio = _normalize_tolerance(max_decrease, quantity)

    # Allowed quantity window
    max_allowed = quantity * (1 + increase_ratio)
    min_allowed = max(0, quantity * (1 - decrease_ratio))

    # ── Step 1: Try full pallets ──────────────────────────────────────
    # Round UP to the next full-pallet boundary
    pallets_up = math.ceil(quantity / pcs_per_plt)
    qty_up = pallets_up * pcs_per_plt
    if qty_up <= max_allowed:
        return qty_up

    # Round DOWN to the previous full-pallet boundary
    pallets_down = quantity // pcs_per_plt
    qty_down = pallets_down * pcs_per_plt
    if qty_down > 0 and qty_down >= min_allowed:
        return qty_down

    # ── Step 2: Mixed packing (pallets + cartons) ─────────────────────
    # Iterate from the most pallets possible down to zero.
    # For each pallet count, find the minimum number of cartons that
    # brings the total into the allowed range.
    max_pallets = int(max_allowed // pcs_per_plt)

    for num_pallets in range(max_pallets, -1, -1):
        pallet_qty = num_pallets * pcs_per_plt

        # How many more pieces are still needed to reach min_allowed?
        remaining_min = max(0, min_allowed - pallet_qty)
        remaining_max = max_allowed - pallet_qty

        if remaining_max < 0:
            continue  # too many pallets already

        # Minimum cartons to cover the remaining gap
        min_cartons = math.ceil(remaining_min / pcs_per_ctn) if remaining_min > 0 else 0
        total = pallet_qty + min_cartons * pcs_per_ctn

        if min_allowed <= total <= max_allowed:
            return int(total)

    # No valid combination found within tolerance
    return None


# ── Data extraction ───────────────────────────────────────────────────

def extract_data_from_csv(csv_path, data_cols):
    """
    Read the CSV and extract the key column (first column) plus *data_cols*.

    Returns a DataFrame with ``[key_col, *found_data_cols]``.
    """
    logger.info("Reading converted data from %s", csv_path)
    df = pd.read_csv(csv_path, dtype=str)

    if df.empty:
        raise Exception("The converted CSV file is empty.")

    key_col = df.columns[0]
    logger.info("Using first column '%s' as the Key.", key_col)

    present_cols = []
    missing_cols = []
    for requested in data_cols:
        resolved = _find_name_icase(df.columns, requested)
        if resolved is not None:
            if resolved not in present_cols:
                present_cols.append(resolved)
        else:
            missing_cols.append(requested)

    if missing_cols:
        logger.warning("Columns not found in CSV: %s", missing_cols)
        logger.warning("Available columns: %s", list(df.columns))

    final_cols = [key_col] + [c for c in present_cols if c != key_col]
    return df[final_cols]


# ── Identifier / text-safe value handling ────────────────────────────

def _is_identifier_code(value):
    """Detect an identifier-like string that must be preserved as text.

    Matches purely-numeric strings that either:

    * start with a leading ``0`` (e.g. ``"000461200000102"``), or
    * exceed 15 digits (where ``float`` loses precision — Excel only
      stores 15 significant digits in numeric cells).
    """
    if not isinstance(value, str):
        return False
    s = value.strip()
    if not s or not s.isdigit():
        return False
    if s.startswith("0") and len(s) > 1:
        return True
    return len(s) > 15


def _canonicalize_excel_value(cell_value):
    """Return a canonical string form for an openpyxl cell value.

    Integer-valued floats like ``461200000102.0`` are stored by Excel as
    numbers, but CSV keys typically arrive as plain digit strings
    (``"461200000102"``).  Canonicalising both sides makes the lookup
    match regardless of the original storage type.
    """
    if cell_value is None:
        return None
    if isinstance(cell_value, float) and cell_value.is_integer():
        return str(int(cell_value))
    if isinstance(cell_value, (int, float)):
        return str(cell_value)
    return str(cell_value).strip()


def _coerce_value_for_excel(raw):
    """Prepare a raw CSV value for writing to a cell.

    Returns a ``(value, force_text)`` pair:

    * ``force_text=True``  — the value is an identifier-like digit string
      and the caller should set ``cell.number_format = '@'`` before
      assignment so Excel doesn't strip leading zeros or convert to
      scientific notation.
    * ``force_text=False`` — the value has been parsed as a number (when
      possible) or passed through as free text; the destination cell's
      existing number format is left untouched.
    """
    if raw is None:
        return "", False
    # Guard against pandas NaN (float('nan') != itself)
    if isinstance(raw, float) and raw != raw:
        return "", False
    if isinstance(raw, str):
        stripped = raw.strip()
        if not stripped:
            return "", False
        if _is_identifier_code(stripped):
            return stripped, True
    return smart_number_convert(raw), False


# ── Excel lookup (read once) ─────────────────────────────────────────

def _build_excel_lookup(excel_path):
    """
    Read the Excel file *once* and return two dicts:

    * ``row_lookup``: ``canonical_cell_value → first 0-based row index``
      Keys are **case-sensitive** (used for matching data values).
    * ``col_lookup``: ``lowercased_canonical_value → first 0-based column index``
      Keys are **lowercased** so all header lookups are case-insensitive.

    Duplicate values record only the **first** (top-left) occurrence.
    Numeric cells are canonicalised via :func:`_canonicalize_excel_value`
    so that an integer stored as ``461200000102`` matches a CSV string
    key of the same value (or its leading-zero-padded variant is handled
    consistently).
    """
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active

    row_lookup = {}
    col_lookup = {}

    try:
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            for c_idx, raw_cell in enumerate(row):
                val = _canonicalize_excel_value(raw_cell)
                if not val:
                    continue
                if val not in row_lookup:
                    row_lookup[val] = r_idx
                val_lower = val.lower()
                if val_lower not in col_lookup:
                    col_lookup[val_lower] = c_idx
    finally:
        wb.close()

    return row_lookup, col_lookup


# ── Case-insensitive column lookup ────────────────────────────────────

_PCS_CTN_COL = "Pcs/Ctn"
_PCS_PLT_COL = "Pcs/Plt"


def _find_name_icase(names, target):
    """Find *target* in *names* using case-insensitive matching.

    Returns the matched original name or ``None``.
    """
    target_lower = target.lower()
    for name in names:
        if name.lower() == target_lower:
            return name
    return None



def _values_equal(existing, new_val):
    """Compare an existing Excel cell value with a new value to insert.

    Handles type mismatches: e.g. int 504 vs float 504.0,
    string "504" vs float 504.0, None vs empty string, etc.
    """
    if existing is None and new_val is None:
        return True
    if existing is None or new_val is None:
        # Treat None and empty string as equal
        if existing is None and isinstance(new_val, str) and new_val == "":
            return True
        if new_val is None and isinstance(existing, str) and existing == "":
            return True
        return False

    # Both are numeric → compare as floats
    if isinstance(existing, (int, float)) and isinstance(new_val, (int, float)):
        return float(existing) == float(new_val)

    # Try numeric comparison for mixed types (e.g. str vs float)
    try:
        return float(existing) == float(new_val)
    except (ValueError, TypeError):
        pass

    # Fall back to stripped string comparison
    return str(existing).strip() == str(new_val).strip()


# ── Public API ────────────────────────────────────────────────────────

def process_insert(csv_path, excel_path, csv_cols, excel_cols,
                   output_path=None, flagged_cells=None,
                   qty_csv_col=None,
                   qty_increase_ratio=0.15, qty_decrease_ratio=0.15):
    """
    Insert columns from a converted CSV into an Excel template.

    The **first column** of the CSV is used as the lookup key.  For every
    row the engine finds the key value in the Excel sheet, then writes the
    requested *csv_cols* values into the corresponding *excel_cols*.

    All updated cells are highlighted **red**.  Cells flagged by validation
    (warnings) are highlighted **yellow** instead.

    When *qty_csv_col* is provided, the engine looks for columns named
    ``Pcs/Ctn`` and ``Pcs/Plt`` (case-insensitive) in the Excel template,
    passes them together with the original quantity to
    ``calculate_new_quantity()``, and writes the result as
    ``"original/calculated"`` (e.g. ``"50/60"``).

    Args:
        csv_path:           Path to the source CSV (output of a converter).
        excel_path:         Path to the Excel template.
        csv_cols:           List of CSV column names to read.
        excel_cols:         List of Excel column headers to write into.
        output_path:        Where to save (default: ``*_merged.xlsx``).
        flagged_cells:      Optional set of ``(key_value, csv_column_name)``
                            tuples.  Matching cells get a yellow fill.
        qty_csv_col:        CSV column name that holds the quantity.

    Returns:
        Path to the saved output file.
    """
    if flagged_cells is None:
        flagged_cells = set()
    logger.info("=== Starting Insert Process ===")
    logger.info("CSV Path:   %s", csv_path)
    logger.info("Excel Path: %s", excel_path)
    logger.info("Mapping:    %s → %s", csv_cols, excel_cols)

    # ── 1. Extract data from CSV ──────────────────────────────────────
    try:
        logger.info("--- Step 1: Extracting Data from CSV ---")
        df_data = extract_data_from_csv(csv_path, csv_cols)
        logger.info("Extracted %d rows.", len(df_data))
        if not df_data.empty:
            logger.debug("Sample keys: %s", df_data.iloc[:3, 0].tolist())
    except Exception as e:
        logger.error("FAILED at Step 1 (Data Extraction): %s", e)
        raise

    # ── 2. Load Excel workbook ────────────────────────────────────────
    try:
        logger.info("--- Step 2: Loading Excel File ---")
        wb = load_workbook(excel_path)
        ws = wb.active
        logger.info("Excel file loaded successfully.")
    except Exception as e:
        logger.error("FAILED at Step 2 (Loading Excel): %s", e)
        raise

    # ── 3. Build lookup maps (single Excel read) ──────────────────────
    logger.info("--- Step 3: Building Excel Lookup Maps ---")
    row_lookup, col_lookup = _build_excel_lookup(excel_path)
    logger.info("Excel lookup built — %d unique cell values, %d unique column headers",
                len(row_lookup), len(col_lookup))

    mapping_pairs = []
    for req_csv_col, req_excel_col in zip(csv_cols, excel_cols):
        resolved_csv_col = _find_name_icase(df_data.columns, req_csv_col)
        if resolved_csv_col is None:
            logger.warning("  CSV column '%s' NOT found (case-insensitive).",
                           req_csv_col)
            continue

        excel_col_idx = col_lookup.get(req_excel_col.lower())
        if excel_col_idx is None:
            logger.warning("  Excel column '%s' NOT found (case-insensitive).",
                           req_excel_col)
            continue

        mapping_pairs.append((resolved_csv_col, excel_col_idx + 1))
        logger.info("  Lookup OK: CSV '%s' → Excel '%s' (column index %d)",
                     resolved_csv_col, req_excel_col, excel_col_idx + 1)

    if not mapping_pairs:
        missing = [e for _, e in zip(csv_cols, excel_cols)
                   if e.lower() not in col_lookup]
        raise ValueError(
            f"None of the Excel columns could be found: {missing}. "
            f"No data was inserted.")

    # ── 3b. Resolve quantity-recalculation columns (optional) ─────────
    qty_recalc_enabled = qty_csv_col is not None
    qty_recalc_requested = qty_csv_col is not None
    resolved_qty_csv_col = None
    pcs_ctn_col_idx = None
    pcs_plt_col_idx = None

    if qty_recalc_enabled:
        resolved_qty_csv_col = _find_name_icase(df_data.columns, qty_csv_col)
        if resolved_qty_csv_col is None:
            logger.warning("  Quantity CSV column '%s' NOT found "
                           "(case-insensitive) — quantity recalculation "
                           "disabled.", qty_csv_col)
            qty_recalc_enabled = False

        ctn_idx = col_lookup.get(_PCS_CTN_COL.lower())
        plt_idx = col_lookup.get(_PCS_PLT_COL.lower())

        if ctn_idx is not None:
            pcs_ctn_col_idx = ctn_idx + 1  # 1-indexed
            logger.info("  Pcs/Ctn column found at index %d",
                        pcs_ctn_col_idx)
        else:
            logger.warning("  '%s' column NOT found (case-insensitive) — "
                           "quantity recalculation disabled.", _PCS_CTN_COL)
            qty_recalc_enabled = False

        if plt_idx is not None:
            pcs_plt_col_idx = plt_idx + 1  # 1-indexed
            logger.info("  Pcs/Plt column found at index %d",
                        pcs_plt_col_idx)
        else:
            logger.warning("  '%s' column NOT found (case-insensitive) — "
                           "quantity recalculation disabled.", _PCS_PLT_COL)
            qty_recalc_enabled = False

    # ── 4. Map CSV rows → Excel rows and update cells ─────────────────
    key_col_name = df_data.columns[0]
    total = len(df_data)
    updates = 0
    red_count = 0
    yellow_count = 0

    not_found_keys = []
    skipped_count = 0
    logger.info("--- Step 4: Updating %d Items ---", total)
    for idx, (_, row) in enumerate(df_data.iterrows(), 1):
        raw_key = row[key_col_name]
        # Canonicalise the lookup key the same way the Excel side was
        # canonicalised so that identifier codes like "000461200000102"
        # match whether the Excel template stored them as text or as a
        # number (which openpyxl reads back as a float).
        item_val = _canonicalize_excel_value(raw_key) or ""
        if not item_val:
            continue

        if idx <= 5 or idx % 10 == 0:
            logger.debug("Processing %d/%d: '%s'", idx, total, item_val)

        if item_val not in row_lookup:
            not_found_keys.append(item_val)
            continue

        row_idx = row_lookup[item_val]
        target_row = row_idx + 1  # openpyxl is 1-indexed

        for c_col, target_col in mapping_pairs:
            val, force_text = _coerce_value_for_excel(row.get(c_col, ""))
            original_val = val  # preserve CSV value before recalculation

            # Quantity recalculation: format as "original/calculated"
            if (qty_recalc_enabled and c_col == resolved_qty_csv_col
                    and pcs_ctn_col_idx and pcs_plt_col_idx
                    and not force_text):
                pcs_ctn_val = ws.cell(
                    row=target_row, column=pcs_ctn_col_idx).value
                pcs_plt_val = ws.cell(
                    row=target_row, column=pcs_plt_col_idx).value
                pcs_ctn = smart_number_convert(pcs_ctn_val)
                pcs_plt = smart_number_convert(pcs_plt_val)

                if (isinstance(pcs_ctn, (int, float))
                        and isinstance(pcs_plt, (int, float))):
                    new_qty = calculate_new_quantity(
                        val, pcs_ctn, pcs_plt,
                        max_increase=qty_increase_ratio,
                        max_decrease=qty_decrease_ratio)
                    if new_qty is None:
                        val = f"{val}/N\u00b7A"
                        force_text = True  # compound string → text cell
                        logger.info(
                            "  Qty recalc for '%s': Pcs/Ctn=%s Pcs/Plt=%s → %s "
                            "(no valid packing within tolerance)",
                            item_val, pcs_ctn, pcs_plt, val)
                    elif not _values_equal(val, new_qty):
                        val = f"{val}/{new_qty}"
                        force_text = True  # compound string → text cell
                        logger.info(
                            "  Qty recalc for '%s': Pcs/Ctn=%s Pcs/Plt=%s → %s",
                            item_val, pcs_ctn, pcs_plt, val)

            # Was the value modified from what the CSV originally had?
            value_was_modified = not _values_equal(original_val, val)

            try:
                cell = ws.cell(row=target_row, column=target_col)
                existing_raw = cell.value
                existing = smart_number_convert(existing_raw)
                # Skip if the value is unchanged
                if _values_equal(existing, val):
                    skipped_count += 1
                    continue

                # Force every inserted cell to Times New Roman 9 while
                # preserving any other font decorations (bold, italic,
                # underline, color) the template had on that cell.
                # Alignment / border / fill are left untouched by this
                # assignment.  number_format is overridden only when the
                # value is an identifier-like string so Excel renders it
                # verbatim.
                prev = cell.font
                cell.font = Font(
                    name=_INSERT_FONT_NAME,
                    size=_INSERT_FONT_SIZE,
                    bold=prev.bold,
                    italic=prev.italic,
                    underline=prev.underline,
                    strike=prev.strike,
                    color=prev.color,
                    vertAlign=prev.vertAlign,
                )
                if force_text:
                    cell.number_format = _TEXT_FORMAT
                cell.value = val

                # Yellow = validation warning
                # Red = value was modified from original CSV value
                # No fill = straight passthrough from CSV
                if (item_val, c_col) in flagged_cells:
                    cell.fill = _YELLOW_FILL
                    yellow_count += 1
                    logger.info("  [YELLOW] '%s' col '%s' = %s (flagged by validation)",
                                item_val, c_col, val)
                elif value_was_modified:
                    cell.fill = _RED_FILL
                    red_count += 1
                    logger.info("  [RED] '%s' col '%s' = %s (modified from CSV %s)",
                                item_val, c_col, val, original_val)
            except Exception as e:
                logger.error("  Error writing (%d, %d): %s",
                             target_row, target_col, e)

        updates += 1

    logger.info("Cell highlights: %d red (updated), %d yellow (flagged), %d skipped (unchanged)",
                red_count, yellow_count, skipped_count)

    if not_found_keys:
        logger.info("%d key(s) NOT found in Excel: %s", len(not_found_keys), not_found_keys)
    else:
        logger.info("All %d CSV key(s) found in Excel.", total)

    # Track columns that could not be resolved
    missing_excel_cols = [
        req_excel_col
        for req_csv_col, req_excel_col in zip(csv_cols, excel_cols)
        if req_excel_col.lower() not in col_lookup
    ]

    insert_stats = {
        "total_csv_rows": total,
        "rows_matched": updates,
        "rows_not_found": len(not_found_keys),
        "red_cells": red_count,
        "yellow_cells": yellow_count,
        "skipped_cells": skipped_count,
        "missing_columns": missing_excel_cols,
        "qty_recalc_disabled": qty_recalc_requested and not qty_recalc_enabled,
    }

    if updates == 0:
        msg = "No matches found. No output file was generated."
        logger.warning(msg)
        raise NoMatchesFoundError(msg, stats=insert_stats)

    # ── 5. Save ───────────────────────────────────────────────────────
    if not output_path:
        output_path = excel_path.replace(".xlsx", "_merged.xlsx")

    logger.info("--- Step 5: Saving to %s ---", output_path)
    try:
        wb.save(output_path)
        logger.info("=== Insert complete: %d/%d rows matched, "
                    "%d column mapping(s), %d red / %d yellow cell(s) ===",
                    updates, total, len(mapping_pairs), red_count, yellow_count)
    except Exception as e:
        logger.error("FAILED at Step 5 (Saving): %s", e)
        raise

    insert_stats["output_path"] = output_path
    return insert_stats


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    # Quick manual test
    import sys
    if len(sys.argv) >= 5:
        csv_p = sys.argv[1]
        excel_p = sys.argv[2]
        c_cols = [c.strip() for c in sys.argv[3].split(",")]
        e_cols = [c.strip() for c in sys.argv[4].split(",")]
        out = sys.argv[5] if len(sys.argv) > 5 else None
        process_insert(csv_p, excel_p, c_cols, e_cols, output_path=out)
    else:
        logger.info("Usage: python -m engine.insert <csv> <excel> <csv_cols> <excel_cols> [output]")
