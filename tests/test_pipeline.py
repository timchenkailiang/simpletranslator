"""Tests for engine.pipeline — end-to-end pipeline orchestration."""

import csv
import os
import sys
import tempfile
import types
import unittest
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from engine.pipeline import (
    run_full_pipeline, run_extract_only,
    _resolve_col_by_position, _resolve_validation_rules,
)
from engine.insert import NoMatchesFoundError


# ── Helpers ───────────────────────────────────────────────────────────

def _make_csv(path, header, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in rows:
            w.writerow(row)


def _make_excel(path, header, rows):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _fake_converter(csv_header, csv_rows, format_name="TestFmt",
                    validation_rules=None):
    """Build a fake converter module that writes a fixed CSV on process_file."""
    mod = types.ModuleType("fake_converter")
    mod.FORMAT_NAME = format_name
    mod.COLUMNS = csv_header
    mod.VALIDATION_RULES = validation_rules or {
        "qty": "Quantity", "price": "Price", "amount": "Amount"
    }

    def process_file(input_path, output_path):
        _make_csv(output_path, csv_header, csv_rows)
        return True

    mod.process_file = process_file
    return mod


# ── Pipeline tests ────────────────────────────────────────────────────

class TestRunFullPipeline(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_basic_full_pipeline(self):
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "intermediate.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        # Create a dummy PDF (converter ignores content)
        with open(pdf_path, "w") as f:
            f.write("dummy")

        _make_excel(excel_path,
                    ["Item", "Qty"],
                    [["A001", ""], ["A002", ""]])

        converter = _fake_converter(
            csv_header=["Item", "Quantity", "Price", "Amount"],
            csv_rows=[["A001", "10", "5.0", "50.0"],
                      ["A002", "20", "3.0", "60.0"]],
        )

        result = run_full_pipeline(
            pdf_path, csv_path, excel_path,
            converter, ["Quantity"], ["Qty"],
            output_path=output_path,
        )
        self.assertEqual(result["output_path"], output_path)
        self.assertEqual(result["rows_extracted"], 2)
        self.assertEqual(result["rows_matched"], 2)
        self.assertTrue(os.path.exists(output_path))

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=2).value, 10.0)
        self.assertEqual(ws.cell(row=3, column=2).value, 20.0)

    def test_pipeline_progress_callback(self):
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "intermediate.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        with open(pdf_path, "w") as f:
            f.write("dummy")
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        converter = _fake_converter(
            csv_header=["Item", "Quantity", "Price", "Amount"],
            csv_rows=[["A001", "10", "5.0", "50.0"]],
        )

        progress_log = []

        def on_progress(pct, msg):
            progress_log.append((pct, msg))

        run_full_pipeline(
            pdf_path, csv_path, excel_path,
            converter, ["Quantity"], ["Qty"],
            output_path=output_path,
            on_progress=on_progress,
        )
        # Should have received multiple progress updates ending at 100
        self.assertTrue(len(progress_log) > 0)
        self.assertEqual(progress_log[-1][0], 100)

    def test_pipeline_no_matches_raises(self):
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "intermediate.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        with open(pdf_path, "w") as f:
            f.write("dummy")
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        converter = _fake_converter(
            csv_header=["Item", "Quantity", "Price", "Amount"],
            csv_rows=[["NOPE", "10", "5.0", "50.0"]],
        )

        with self.assertRaises(NoMatchesFoundError) as ctx:
            run_full_pipeline(
                pdf_path, csv_path, excel_path,
                converter, ["Quantity"], ["Qty"],
                output_path=output_path,
            )
        self.assertEqual(ctx.exception.stats["rows_extracted"], 1)
        self.assertEqual(ctx.exception.stats["rows_matched"], 0)

    def test_pipeline_conversion_failure(self):
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "intermediate.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        with open(pdf_path, "w") as f:
            f.write("dummy")
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        # Converter that returns False and writes no CSV
        mod = types.ModuleType("fail_converter")
        mod.FORMAT_NAME = "Fail"
        mod.COLUMNS = []
        mod.VALIDATION_RULES = {"qty": "Quantity", "price": "Price", "amount": "Amount"}
        mod.process_file = lambda inp, out: False

        with self.assertRaises(RuntimeError):
            run_full_pipeline(
                pdf_path, csv_path, excel_path,
                mod, ["Quantity"], ["Qty"],
                output_path=output_path,
            )

    def test_auto_detect_qty_col(self):
        """Pipeline auto-detects qty_csv_col from VALIDATION_RULES."""
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "intermediate.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        with open(pdf_path, "w") as f:
            f.write("dummy")
        _make_excel(excel_path,
                    ["Item", "Qty", "Pcs/Ctn", "Pcs/Plt"],
                    [["A001", "", "10", "50"]])

        converter = _fake_converter(
            csv_header=["Item", "Quantity", "Price", "Amount"],
            csv_rows=[["A001", "100", "5.0", "500.0"]],
        )

        # Don't pass qty_csv_col — should be auto-detected
        result = run_full_pipeline(
            pdf_path, csv_path, excel_path,
            converter, ["Quantity"], ["Qty"],
            output_path=output_path,
        )
        self.assertTrue(os.path.exists(result["output_path"]))


class TestRunExtractOnly(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_extract_only(self):
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "output.csv")

        with open(pdf_path, "w") as f:
            f.write("dummy")

        converter = _fake_converter(
            csv_header=["Item", "Quantity", "Price", "Amount"],
            csv_rows=[["A001", "10", "5.0", "50.0"]],
        )

        result_csv, flagged = run_extract_only(
            pdf_path, csv_path, converter,
        )
        self.assertEqual(result_csv, csv_path)
        self.assertTrue(os.path.exists(csv_path))
        self.assertIsInstance(flagged, set)

    def test_extract_only_failure(self):
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "output.csv")

        with open(pdf_path, "w") as f:
            f.write("dummy")

        mod = types.ModuleType("fail_converter")
        mod.FORMAT_NAME = "Fail"
        mod.COLUMNS = []
        mod.VALIDATION_RULES = {}
        mod.process_file = lambda inp, out: False

        with self.assertRaises(RuntimeError):
            run_extract_only(pdf_path, csv_path, mod)


# ── Column resolution helper tests ───────────────────────────────────

class TestResolveColByPosition(unittest.TestCase):
    """_resolve_col_by_position maps English metadata names to localised
    CSV headers by matching the index in COLUMNS."""

    def test_english_headers_unchanged(self):
        actual = ["Item", "Quantity", "Price"]
        columns = ["Item", "Quantity", "Price"]
        self.assertEqual(
            _resolve_col_by_position("Quantity", actual, columns), "Quantity")

    def test_spanish_headers_resolved(self):
        actual = ["Artículo", "Cantidad", "Precio"]
        columns = ["Item", "Description", "Quantity", "Price"]
        # "Quantity" is at index 2 in COLUMNS, actual[2] = "Precio"
        # Wait — that's wrong. Let me match the Globe converter:
        # COLUMNS = ["Item", "Description", "Quantity", "Price", "Amount", "Delivery"]
        # actual  = ["Artículo", "Denominación", "Cantidad", "Precio (USD)", "Importe", "F. Entrega"]
        actual = ["Artículo", "Denominación", "Cantidad", "Precio (USD)",
                  "Importe", "F. Entrega"]
        columns = ["Item", "Description", "Quantity", "Price",
                   "Amount", "Delivery"]
        self.assertEqual(
            _resolve_col_by_position("Quantity", actual, columns), "Cantidad")
        self.assertEqual(
            _resolve_col_by_position("Price", actual, columns), "Precio (USD)")
        self.assertEqual(
            _resolve_col_by_position("Amount", actual, columns), "Importe")

    def test_chinese_headers_resolved(self):
        actual = ["物品", "描述", "数量", "价格", "金额", "交期"]
        columns = ["Item", "Description", "Quantity", "Price",
                   "Amount", "Delivery"]
        self.assertEqual(
            _resolve_col_by_position("Quantity", actual, columns), "数量")

    def test_already_present_unchanged(self):
        actual = ["Item", "Cantidad", "Precio"]
        columns = ["Item", "Quantity", "Price"]
        # "Cantidad" is already in actual — but we're resolving "Quantity"
        # which is NOT in actual, so it resolves by position
        self.assertEqual(
            _resolve_col_by_position("Quantity", actual, columns), "Cantidad")

    def test_not_in_columns_returns_original(self):
        actual = ["Item", "Cantidad"]
        columns = ["Item", "Description"]
        self.assertEqual(
            _resolve_col_by_position("Quantity", actual, columns), "Quantity")

    def test_none_col_returns_none(self):
        self.assertIsNone(_resolve_col_by_position(None, ["A"], ["A"]))

    def test_no_columns_meta_returns_original(self):
        self.assertEqual(
            _resolve_col_by_position("Quantity", ["A", "B"], None), "Quantity")

    def test_case_insensitive_match(self):
        actual = ["item", "cantidad"]
        columns = ["Item", "Quantity"]
        self.assertEqual(
            _resolve_col_by_position("quantity", actual, columns), "cantidad")


class TestResolveValidationRules(unittest.TestCase):
    """_resolve_validation_rules maps all rules at once."""

    def test_spanish_rules_resolved(self):
        rules = {"qty": "Quantity", "price": "Price", "amount": "Amount"}
        actual = ["Artículo", "Denominación", "Cantidad", "Precio (USD)",
                  "Importe", "F. Entrega"]
        columns = ["Item", "Description", "Quantity", "Price",
                   "Amount", "Delivery"]
        resolved = _resolve_validation_rules(rules, actual, columns)
        self.assertEqual(resolved["qty"], "Cantidad")
        self.assertEqual(resolved["price"], "Precio (USD)")
        self.assertEqual(resolved["amount"], "Importe")

    def test_english_rules_unchanged(self):
        rules = {"qty": "Quantity", "price": "Price", "amount": "Amount"}
        actual = ["Item", "Description", "Quantity", "Price",
                  "Amount", "Delivery"]
        columns = ["Item", "Description", "Quantity", "Price",
                   "Amount", "Delivery"]
        resolved = _resolve_validation_rules(rules, actual, columns)
        self.assertEqual(resolved["qty"], "Quantity")

    def test_none_rules_returns_none(self):
        self.assertIsNone(_resolve_validation_rules(None, ["A"], ["A"]))

    def test_none_value_preserved(self):
        rules = {"qty": "Quantity", "amount": None}
        actual = ["Artículo", "Cantidad"]
        columns = ["Item", "Quantity"]
        resolved = _resolve_validation_rules(rules, actual, columns)
        self.assertIsNone(resolved["amount"])


# ── Integration: localised CSV + full pipeline ────────────────────────

class TestLocalisedPipeline(unittest.TestCase):
    """End-to-end tests with localised (non-English) CSV headers."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_localised_converter(self, localised_headers):
        """Fake converter that writes localised headers but has English COLUMNS
        (matching how real converters like Globe work)."""
        english_columns = ["Item", "Description", "Quantity", "Price",
                           "Amount", "Delivery"]
        mod = types.ModuleType("localised_converter")
        mod.FORMAT_NAME = "Globe"
        mod.COLUMNS = english_columns
        mod.VALIDATION_RULES = {"qty": "Quantity", "price": "Price",
                                "amount": "Amount"}

        def process_file(input_path, output_path):
            _make_csv(output_path, localised_headers,
                      [["000461200000102", "ESCUADRA FIJA",
                        "504", "0.818", "412.27", "29/12/2025"]])
            return True

        mod.process_file = process_file
        return mod

    def test_spanish_qty_recalc_enabled(self):
        """Qty recalculation works when CSV has Spanish headers."""
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "intermediate.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        with open(pdf_path, "w") as f:
            f.write("dummy")
        _make_excel(excel_path,
                    ["Key", "Qty", "Pcs/Ctn", "Pcs/Plt"],
                    [["000461200000102", "", 10, 50]])

        spanish_headers = ["Artículo", "Denominación", "Cantidad",
                           "Precio (USD)", "Importe", "F. Entrega"]
        converter = self._make_localised_converter(spanish_headers)

        result = run_full_pipeline(
            pdf_path, csv_path, excel_path,
            converter, ["Cantidad"], ["Qty"],
            output_path=output_path,
        )

        # Qty recalc should have fired: 504 → 504.0/500
        # (Pcs/Ctn=10, Pcs/Plt=50 → round down to 500)
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        cell_val = ws.cell(row=2, column=2).value
        self.assertIsNotNone(cell_val)
        self.assertIn("/", str(cell_val))  # "504.0/500" format
        self.assertFalse(result["qty_recalc_disabled"])

    def test_spanish_validation_rules_resolved(self):
        """Validation checks math even with Spanish column names."""
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "intermediate.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        with open(pdf_path, "w") as f:
            f.write("dummy")
        _make_excel(excel_path, ["Key", "Qty"],
                    [["000461200000102", ""]])

        # Intentional math error: 504 × 0.818 = 412.272, but amount = 999
        bad_converter = types.ModuleType("bad_converter")
        bad_converter.FORMAT_NAME = "Globe"
        bad_converter.COLUMNS = ["Item", "Description", "Quantity", "Price",
                                  "Amount", "Delivery"]
        bad_converter.VALIDATION_RULES = {"qty": "Quantity", "price": "Price",
                                          "amount": "Amount"}
        def bad_process(inp, out):
            _make_csv(out,
                      ["Artículo", "Denominación", "Cantidad",
                       "Precio (USD)", "Importe", "F. Entrega"],
                      [["000461200000102", "ESCUADRA", "504", "0.818",
                        "999", "29/12/2025"]])
            return True
        bad_converter.process_file = bad_process

        result = run_full_pipeline(
            pdf_path, csv_path, excel_path,
            bad_converter, ["Cantidad"], ["Qty"],
            output_path=output_path,
        )
        # Should have flagged the math error
        self.assertGreater(result["cells_flagged"], 0)

    def test_extract_only_spanish_validation(self):
        """run_extract_only also resolves validation rules for localised CSV."""
        pdf_path = os.path.join(self.tmpdir, "dummy.pdf")
        csv_path = os.path.join(self.tmpdir, "output.csv")

        with open(pdf_path, "w") as f:
            f.write("dummy")

        # Math error: 504 × 0.818 ≠ 999
        bad_converter = types.ModuleType("bad_converter")
        bad_converter.FORMAT_NAME = "Globe"
        bad_converter.COLUMNS = ["Item", "Description", "Quantity", "Price",
                                  "Amount", "Delivery"]
        bad_converter.VALIDATION_RULES = {"qty": "Quantity", "price": "Price",
                                          "amount": "Amount"}
        def bad_process(inp, out):
            _make_csv(out,
                      ["Artículo", "Denominación", "Cantidad",
                       "Precio (USD)", "Importe", "F. Entrega"],
                      [["000461200000102", "ESCUADRA", "504", "0.818",
                        "999", "29/12/2025"]])
            return True
        bad_converter.process_file = bad_process

        _, flagged = run_extract_only(pdf_path, csv_path, bad_converter)
        self.assertGreater(len(flagged), 0)


if __name__ == "__main__":
    unittest.main()
