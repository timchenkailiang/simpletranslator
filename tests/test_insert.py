"""Tests for engine.insert — CSV→Excel insertion, lookup, and cell highlighting."""

import csv
import os
import sys
import tempfile
import unittest
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from engine.insert import (
    extract_data_from_csv,
    _build_excel_lookup,
    _find_name_icase,
    _values_equal,
    _is_identifier_code,
    _coerce_value_for_excel,
    _canonicalize_excel_value,
    process_insert,
    calculate_new_quantity,
    NoMatchesFoundError,
)


# ── Helper: write a minimal Excel workbook ────────────────────────────

def _make_excel(path, header, rows):
    """Create a tiny .xlsx with openpyxl for testing."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _make_csv(path, header, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in rows:
            w.writerow(row)


# ── Unit tests ────────────────────────────────────────────────────────

class TestFindNameIcase(unittest.TestCase):
    def test_exact_match(self):
        self.assertEqual(_find_name_icase(["Foo", "Bar"], "Foo"), "Foo")

    def test_case_insensitive(self):
        self.assertEqual(_find_name_icase(["Quantity", "Price"], "quantity"), "Quantity")

    def test_not_found(self):
        self.assertIsNone(_find_name_icase(["A", "B"], "C"))


class TestCalculateNewQuantity(unittest.TestCase):
    def test_exact_pallet_unchanged(self):
        # Quantity already fills exact pallets — no adjustment needed
        self.assertEqual(calculate_new_quantity(100, 10, 50), 100)

    def test_round_up_to_full_pallet(self):
        # 90 pcs, pallet=50 → round up to 100 (2 pallets, +11 %)
        self.assertEqual(calculate_new_quantity(90, 10, 50), 100)

    def test_round_down_to_full_pallet(self):
        # 104 pcs, pallet=50, increase=0 → can't round up,
        # round down to 100 (2 pallets, −3.8 %) within 15 % decrease
        self.assertEqual(
            calculate_new_quantity(104, 10, 50,
                                   max_increase=0.0,
                                   max_decrease=0.15), 100)

    def test_mixed_packing(self):
        # 170 pcs, pallet=100, carton=25, increase max +5 %
        # Full pallet up = 200 (+17.6 %) — too much
        # Full pallet down = 100 (−41 %) — too much with 5 % decrease
        # Mixed: 1 pallet (100) + 3 cartons (75) = 175 — within +5 %? 175/170 ≈ +2.9 % ✓
        self.assertEqual(
            calculate_new_quantity(170, 25, 100,
                                   max_increase=0.05,
                                   max_decrease=0.05), 175)

    def test_no_valid_option_returns_none(self):
        # Very tight tolerance — no valid packing fits
        result = calculate_new_quantity(99, 50, 200,
                                        max_increase=0.0,
                                        max_decrease=0.0)
        self.assertIsNone(result)

    def test_zero_quantity(self):
        self.assertEqual(calculate_new_quantity(0, 10, 50), 0)

    def test_invalid_packing_returns_unchanged(self):
        self.assertEqual(calculate_new_quantity(100, 0, 50), 100)
        self.assertEqual(calculate_new_quantity(100, 10, 0), 100)

    # ── Tolerance input format tests ──────────────────────────────────
    def test_percentage_string(self):
        # "15%" → same as ratio 0.15
        self.assertEqual(
            calculate_new_quantity(90, 10, 50, max_increase="15%"),
            calculate_new_quantity(90, 10, 50, max_increase=0.15))

    def test_percentage_string_with_spaces(self):
        self.assertEqual(
            calculate_new_quantity(90, 10, 50, max_increase=" 15 % "),
            calculate_new_quantity(90, 10, 50, max_increase=0.15))

    def test_exact_piece_count_increase(self):
        # 90 pcs, pallet=50, allow +10 pieces → max 100, rounds to 100
        self.assertEqual(
            calculate_new_quantity(90, 10, 50, max_increase=10), 100)

    def test_exact_piece_count_decrease(self):
        # 104 pcs, pallet=50, allow −5 pcs (≈4.8 %) → min 99,
        # round down to 100 (2 pallets)
        self.assertEqual(
            calculate_new_quantity(104, 10, 50,
                                   max_increase=0, max_decrease=5), 100)

    def test_exact_piece_count_too_small(self):
        # 90 pcs, pallet=50, allow +5/−5 pieces → range [85, 95],
        # can't reach 100; mixed: 1 pallet(50) + 4 cartons(40) = 90
        self.assertEqual(
            calculate_new_quantity(90, 10, 50, max_increase=5, max_decrease=5), 90)


class TestExtractDataFromCsv(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_extracts_key_and_requested_cols(self):
        path = os.path.join(self.tmpdir, "test.csv")
        _make_csv(path,
                  ["Item", "Description", "Quantity", "Price"],
                  [["A1", "Widget", "10", "5.0"],
                   ["A2", "Gadget", "20", "3.0"]])
        df = extract_data_from_csv(path, ["Quantity"])
        self.assertListEqual(list(df.columns), ["Item", "Quantity"])
        self.assertEqual(len(df), 2)

    def test_missing_col_still_works(self):
        path = os.path.join(self.tmpdir, "test.csv")
        _make_csv(path,
                  ["Item", "Quantity"],
                  [["A1", "10"]])
        df = extract_data_from_csv(path, ["Quantity", "NonExistent"])
        self.assertIn("Quantity", df.columns)
        self.assertNotIn("NonExistent", df.columns)

    def test_empty_csv_raises(self):
        path = os.path.join(self.tmpdir, "empty.csv")
        _make_csv(path, ["Item", "Quantity"], [])
        with self.assertRaises(Exception):
            extract_data_from_csv(path, ["Quantity"])


class TestBuildExcelLookup(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_row_and_col_lookup(self):
        path = os.path.join(self.tmpdir, "test.xlsx")
        _make_excel(path,
                    ["Item", "Quantity", "Price"],
                    [["A001", "10", "5.0"],
                     ["A002", "20", "3.0"]])
        row_lookup, col_lookup = _build_excel_lookup(path)

        # Header row cells should be in both lookups
        self.assertIn("item", col_lookup)        # lowercased
        self.assertIn("quantity", col_lookup)
        # Data values should be in row_lookup (case-sensitive)
        self.assertIn("A001", row_lookup)
        self.assertIn("A002", row_lookup)

    def test_first_occurrence_wins(self):
        path = os.path.join(self.tmpdir, "dup.xlsx")
        _make_excel(path,
                    ["Key", "Val"],
                    [["X", "1"],
                     ["X", "2"]])  # duplicate "X"
        row_lookup, _ = _build_excel_lookup(path)
        # First occurrence (row index 1, since row 0 is header) should win
        self.assertEqual(row_lookup["X"], 1)


# ── Integration tests for process_insert ──────────────────────────────

class TestProcessInsert(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_basic_insert(self):
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path,
                  ["Item", "Quantity"],
                  [["A001", "100"],
                   ["A002", "200"]])
        _make_excel(excel_path,
                    ["Item", "Qty"],
                    [["A001", ""],
                     ["A002", ""],
                     ["A003", ""]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path,
        )
        self.assertEqual(result["output_path"], output_path)
        self.assertEqual(result["rows_matched"], 2)
        self.assertEqual(result["rows_not_found"], 0)
        self.assertTrue(os.path.exists(output_path))

        # Verify values were written
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        # Row 2 = A001, Col 2 = Qty
        self.assertEqual(ws.cell(row=2, column=2).value, 100.0)
        self.assertEqual(ws.cell(row=3, column=2).value, 200.0)
        # A003 was not in CSV, should remain empty
        self.assertIsNone(ws.cell(row=4, column=2).value)

    def test_no_fill_for_passthrough(self):
        """Passthrough values (no recalculation) should NOT get red fill."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "50"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        process_insert(csv_path, excel_path,
                       csv_cols=["Quantity"], excel_cols=["Qty"],
                       output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        self.assertEqual(cell.value, 50.0)
        self.assertIn(cell.fill.patternType, (None, "none"))

    def test_red_fill_for_recalculated(self):
        """Red fill only when quantity recalculation modifies the value."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "90"]])
        # Pcs/Ctn=10, Pcs/Plt=50 → recalc 90 → 100 → written as "90.0/100"
        _make_excel(excel_path,
                    ["Item", "Qty", "Pcs/Ctn", "Pcs/Plt"],
                    [["A001", "", 10, 50]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path,
            qty_csv_col="Quantity")

        self.assertEqual(result["red_cells"], 1)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        fill = ws.cell(row=2, column=2).fill
        self.assertEqual(fill.start_color.rgb, "00FFC7CE")

    def test_yellow_fill_for_flagged(self):
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "50"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        flagged = {("A001", "Quantity")}
        process_insert(csv_path, excel_path,
                       csv_cols=["Quantity"], excel_cols=["Qty"],
                       output_path=output_path,
                       flagged_cells=flagged)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        fill = ws.cell(row=2, column=2).fill
        self.assertEqual(fill.start_color.rgb, "00FFFF00")

    def test_no_matches_raises(self):
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["NOPE", "10"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        with self.assertRaises(NoMatchesFoundError) as ctx:
            process_insert(csv_path, excel_path,
                           csv_cols=["Quantity"], excel_cols=["Qty"],
                           output_path=output_path)
        self.assertEqual(ctx.exception.stats["rows_matched"], 0)
        self.assertEqual(ctx.exception.stats["rows_not_found"], 1)

    def test_multiple_column_mapping(self):
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path,
                  ["Item", "Quantity", "Price"],
                  [["A001", "10", "5.5"]])
        _make_excel(excel_path,
                    ["Item", "Qty", "UnitPrice"],
                    [["A001", "", ""]])

        result = process_insert(csv_path, excel_path,
                       csv_cols=["Quantity", "Price"],
                       excel_cols=["Qty", "UnitPrice"],
                       output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=2).value, 10.0)
        self.assertEqual(ws.cell(row=2, column=3).value, 5.5)
        self.assertEqual(result["missing_columns"], [])

    def test_all_excel_cols_not_found_raises(self):
        """ValueError when none of the requested Excel columns exist."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "10"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        with self.assertRaises(ValueError):
            process_insert(csv_path, excel_path,
                           csv_cols=["Quantity"],
                           excel_cols=["NonExistentCol"],
                           output_path=output_path)

    def test_partial_excel_col_missing_reported(self):
        """Missing columns tracked in stats when some columns resolve."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path,
                  ["Item", "Quantity", "Price"],
                  [["A001", "10", "5.0"]])
        _make_excel(excel_path,
                    ["Item", "Qty"],
                    [["A001", ""]])

        result = process_insert(csv_path, excel_path,
                       csv_cols=["Quantity", "Price"],
                       excel_cols=["Qty", "BadCol"],
                       output_path=output_path)
        self.assertIn("BadCol", result["missing_columns"])
        self.assertEqual(result["rows_matched"], 1)


# ── _values_equal unit tests ──────────────────────────────────────────

class TestValuesEqual(unittest.TestCase):
    """Ensure _values_equal handles the type mismatches between
    openpyxl cell values and smart_number_convert output."""

    # Numeric identity
    def test_both_none(self):
        self.assertTrue(_values_equal(None, None))

    def test_none_vs_empty_string(self):
        self.assertTrue(_values_equal(None, ""))
        self.assertTrue(_values_equal("", None))

    def test_none_vs_value(self):
        self.assertFalse(_values_equal(None, 100))
        self.assertFalse(_values_equal(100, None))

    def test_int_vs_float_same(self):
        self.assertTrue(_values_equal(504, 504.0))

    def test_float_vs_float_same(self):
        self.assertTrue(_values_equal(5.5, 5.5))

    def test_int_vs_int_same(self):
        self.assertTrue(_values_equal(100, 100))

    def test_int_vs_float_different(self):
        self.assertFalse(_values_equal(504, 505.0))

    # String vs numeric (typical CSV→Excel mismatch)
    def test_str_vs_float(self):
        self.assertTrue(_values_equal("504", 504.0))

    def test_float_vs_str(self):
        self.assertTrue(_values_equal(504.0, "504"))

    def test_str_vs_int(self):
        self.assertTrue(_values_equal("200", 200))

    def test_str_float_vs_float(self):
        self.assertTrue(_values_equal("5.50", 5.5))

    def test_str_vs_str_same(self):
        self.assertTrue(_values_equal("hello", "hello"))

    def test_str_vs_str_different(self):
        self.assertFalse(_values_equal("hello", "world"))

    def test_str_whitespace_stripped(self):
        self.assertTrue(_values_equal("  hello ", "hello"))

    # Non-numeric strings should not match numbers
    def test_str_non_numeric_vs_float(self):
        self.assertFalse(_values_equal("abc", 100.0))


# ── Skip-unchanged + red/yellow highlight integration tests ───────────

class TestInsertSkipUnchanged(unittest.TestCase):
    """process_insert should skip cells whose value already matches,
    and only red-highlight cells that were actually changed."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_skip_when_value_already_matches_int(self):
        """Excel cell already has 100 (int), CSV has '100' → skip, no red."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "100"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", 100]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path)

        self.assertEqual(result["red_cells"], 0)
        self.assertEqual(result["skipped_cells"], 1)

    def test_skip_when_value_already_matches_float(self):
        """Excel cell has 5.5 (float), CSV has '5.5' → skip."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Price"], [["A001", "5.5"]])
        _make_excel(excel_path, ["Item", "UnitPrice"], [["A001", 5.5]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Price"], excel_cols=["UnitPrice"],
            output_path=output_path)

        self.assertEqual(result["red_cells"], 0)
        self.assertEqual(result["skipped_cells"], 1)

    def test_skip_when_value_matches_string(self):
        """Excel cell has text 'Hello', CSV has 'Hello' → skip."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Note"], [["A001", "Hello"]])
        _make_excel(excel_path, ["Item", "Comment"], [["A001", "Hello"]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Note"], excel_cols=["Comment"],
            output_path=output_path)

        self.assertEqual(result["red_cells"], 0)
        self.assertEqual(result["skipped_cells"], 1)

    def test_no_fill_when_value_different_passthrough(self):
        """Excel cell has 50, CSV has '100', no recalc → write but no red."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "100"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", 50]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path)

        self.assertEqual(result["red_cells"], 0)
        self.assertEqual(result["skipped_cells"], 0)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=2).value, 100.0)
        # No fill — value was not recalculated
        self.assertIn(ws.cell(row=2, column=2).fill.patternType, (None, "none"))

    def test_no_red_when_cell_empty_passthrough(self):
        """Excel cell is empty, CSV has value, no recalc → write, no red."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "100"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path)

        self.assertEqual(result["red_cells"], 0)
        self.assertEqual(result["skipped_cells"], 0)

    def test_mixed_skip_and_write(self):
        """Two rows: one unchanged (skip), one different (write, no red)."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path,
                  ["Item", "Quantity"],
                  [["A001", "100"], ["A002", "200"]])
        _make_excel(excel_path,
                    ["Item", "Qty"],
                    [["A001", 100], ["A002", 999]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path)

        # No recalculation → no red even though value changed
        self.assertEqual(result["red_cells"], 0)
        self.assertEqual(result["skipped_cells"], 1)

    def test_red_only_for_recalculated_value(self):
        """Two rows with recalc: one recalculated (red), one unchanged (skip)."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        # A001: qty=90, recalc → 100 (changed), A002: qty=100, recalc → 100 (unchanged)
        _make_csv(csv_path,
                  ["Item", "Quantity"],
                  [["A001", "90"], ["A002", "100"]])
        _make_excel(excel_path,
                    ["Item", "Qty", "Pcs/Ctn", "Pcs/Plt"],
                    [["A001", "", 10, 50],
                     ["A002", "", 10, 50]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path,
            qty_csv_col="Quantity")

        # A001: 90→"90.0/100" (recalculated, red)
        # A002: 100→100 (unchanged after recalc, no red)
        self.assertEqual(result["red_cells"], 1)
        self.assertEqual(result["skipped_cells"], 0)

    def test_no_fill_on_skipped_cell(self):
        """Skipped cells should NOT have red or yellow fill."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "100"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", 100]])

        process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        fill = ws.cell(row=2, column=2).fill
        # No fill applied — default pattern is "none" or rgb is "00000000"
        self.assertIn(fill.patternType, (None, "none"))


# ── qty_recalc_disabled stat tests ────────────────────────────────────

class TestQtyRecalcDisabledStat(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_qty_recalc_disabled_when_pcs_cols_missing(self):
        """When qty_csv_col is set but Pcs/Ctn and Pcs/Plt are absent."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "100"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path,
            qty_csv_col="Quantity")

        self.assertTrue(result["qty_recalc_disabled"])

    def test_qty_recalc_not_disabled_when_not_requested(self):
        """When qty_csv_col is None, qty_recalc_disabled should be False."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "100"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path)

        self.assertFalse(result["qty_recalc_disabled"])

    def test_qty_recalc_enabled_when_pcs_cols_present(self):
        """When Pcs/Ctn and Pcs/Plt exist, recalc should stay enabled."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "100"]])
        _make_excel(excel_path,
                    ["Item", "Qty", "Pcs/Ctn", "Pcs/Plt"],
                    [["A001", "", 10, 50]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path,
            qty_csv_col="Quantity")

        self.assertFalse(result["qty_recalc_disabled"])


# ── Identifier / leading-zero preservation ───────────────────────────

class TestIsIdentifierCode(unittest.TestCase):
    def test_leading_zero_preserved(self):
        self.assertTrue(_is_identifier_code("000461200000102"))

    def test_leading_zero_short(self):
        self.assertTrue(_is_identifier_code("0123"))

    def test_single_zero_is_not_identifier(self):
        # "0" by itself is just the number zero, not a code.
        self.assertFalse(_is_identifier_code("0"))

    def test_long_digit_string_is_identifier(self):
        # 16-digit number exceeds float precision — treat as text.
        self.assertTrue(_is_identifier_code("1234567890123456"))

    def test_short_number_is_not_identifier(self):
        self.assertFalse(_is_identifier_code("123"))

    def test_non_digit_is_not_identifier(self):
        self.assertFalse(_is_identifier_code("A0001"))
        self.assertFalse(_is_identifier_code(""))
        self.assertFalse(_is_identifier_code("  "))

    def test_non_string_is_not_identifier(self):
        self.assertFalse(_is_identifier_code(0))
        self.assertFalse(_is_identifier_code(None))
        self.assertFalse(_is_identifier_code(100.5))


class TestCoerceValueForExcel(unittest.TestCase):
    def test_leading_zero_returns_text(self):
        val, force_text = _coerce_value_for_excel("000461200000102")
        self.assertEqual(val, "000461200000102")
        self.assertTrue(force_text)

    def test_plain_integer_becomes_number(self):
        val, force_text = _coerce_value_for_excel("100")
        self.assertEqual(val, 100.0)
        self.assertFalse(force_text)

    def test_decimal_becomes_number(self):
        val, force_text = _coerce_value_for_excel("5.5")
        self.assertEqual(val, 5.5)
        self.assertFalse(force_text)

    def test_free_text_passes_through(self):
        val, force_text = _coerce_value_for_excel("Hello")
        self.assertEqual(val, "Hello")
        self.assertFalse(force_text)

    def test_none_returns_blank(self):
        val, force_text = _coerce_value_for_excel(None)
        self.assertEqual(val, "")
        self.assertFalse(force_text)

    def test_nan_returns_blank(self):
        val, force_text = _coerce_value_for_excel(float("nan"))
        self.assertEqual(val, "")
        self.assertFalse(force_text)


class TestCanonicalizeExcelValue(unittest.TestCase):
    def test_integer_valued_float_stripped(self):
        self.assertEqual(_canonicalize_excel_value(461200000102.0),
                         "461200000102")

    def test_non_integer_float_preserved(self):
        self.assertEqual(_canonicalize_excel_value(5.5), "5.5")

    def test_string_stripped(self):
        self.assertEqual(_canonicalize_excel_value("  A001 "), "A001")

    def test_none_returns_none(self):
        self.assertIsNone(_canonicalize_excel_value(None))


class TestLeadingZeroPreservation(unittest.TestCase):
    """End-to-end: product codes with leading zeros stay intact in the
    merged Excel file and do not switch to scientific notation."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_leading_zero_part_number_written_as_text(self):
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        # Scenario: the merge mirrors the product code from the CSV into
        # a "PartCode" column of the Excel template.  The CSV has the
        # code with leading zeros; the Excel template has the key in a
        # plain "Ref" column and a blank "PartCode" column to populate.
        _make_csv(csv_path,
                  ["Ref", "PartCode"],
                  [["A001", "000461200000102"],
                   ["A002", "00123"]])
        _make_excel(excel_path,
                    ["Ref", "PartCode"],
                    [["A001", ""], ["A002", ""]])

        process_insert(
            csv_path, excel_path,
            csv_cols=["PartCode"], excel_cols=["PartCode"],
            output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        cell_a = ws.cell(row=2, column=2)
        cell_b = ws.cell(row=3, column=2)
        self.assertEqual(cell_a.value, "000461200000102")
        self.assertEqual(cell_a.number_format, "@")
        self.assertEqual(cell_b.value, "00123")
        self.assertEqual(cell_b.number_format, "@")

    def test_plain_number_still_written_as_number(self):
        """Regular quantities are still stored as numbers (not text)."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "504"]])
        _make_excel(excel_path, ["Item", "Qty"], [["A001", ""]])

        process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        # Written as a number, not text — and its number_format was NOT
        # overridden, so the template's formatting remains in effect.
        self.assertEqual(cell.value, 504.0)
        self.assertNotEqual(cell.number_format, "@")


class TestNumericKeyLookup(unittest.TestCase):
    """The Excel template may store the key column as a number (which
    openpyxl reads back as a float).  The CSV key will still arrive as a
    plain digit string.  _build_excel_lookup should canonicalise both
    sides so the match succeeds."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_numeric_excel_key_matches_string_csv_key(self):
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        # Excel stores the Ref as a real number 600200107 (no leading
        # zero so nothing is lost).  CSV has the string "600200107".
        _make_csv(csv_path, ["Ref", "Qty"], [["600200107", "1000"]])
        _make_excel(excel_path, ["Ref", "Qty"], [[600200107, ""]])

        result = process_insert(
            csv_path, excel_path,
            csv_cols=["Qty"], excel_cols=["Qty"],
            output_path=output_path)
        self.assertEqual(result["rows_matched"], 1)
        self.assertEqual(result["rows_not_found"], 0)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=2).value, 1000.0)


class TestFormatPreservation(unittest.TestCase):
    """Writing a value should preserve the destination cell's existing
    typography (font, alignment, number format for numeric cells)."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_styled_template(self, path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.append(["Item", "Qty"])
        ws.append(["A001", None])
        # Apply a distinctive font + alignment + number format to the
        # destination cell so we can assert nothing clobbers it.
        cell = ws.cell(row=2, column=2)
        cell.font = Font(name="Times New Roman", size=14, bold=True)
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = "#,##0"
        wb.save(path)

    def test_inserted_cell_is_times_new_roman_9(self):
        """Every inserted cell is stamped Times New Roman size 9; the
        template's alignment / border / number format are preserved."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Item", "Quantity"], [["A001", "2500"]])
        self._make_styled_template(excel_path)

        process_insert(
            csv_path, excel_path,
            csv_cols=["Quantity"], excel_cols=["Qty"],
            output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        self.assertEqual(cell.value, 2500.0)
        # Font is forced to Times New Roman 9, but bold carries over.
        self.assertEqual(cell.font.name, "Times New Roman")
        self.assertEqual(cell.font.size, 9)
        self.assertTrue(cell.font.bold)
        # Alignment + number format survive untouched.
        self.assertEqual(cell.alignment.horizontal, "right")
        self.assertEqual(cell.number_format, "#,##0")

    def test_font_forced_even_when_writing_text_identifier(self):
        """Text-format (identifier) cells are still stamped TNR 9."""
        csv_path = os.path.join(self.tmpdir, "data.csv")
        excel_path = os.path.join(self.tmpdir, "template.xlsx")
        output_path = os.path.join(self.tmpdir, "output.xlsx")

        _make_csv(csv_path, ["Ref", "PartCode"],
                  [["A001", "000461200000102"]])
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.append(["Ref", "PartCode"])
        ws.append(["A001", None])
        target = ws.cell(row=2, column=2)
        target.font = Font(name="Arial", size=12, italic=True)
        target.alignment = Alignment(horizontal="center")
        wb.save(excel_path)

        process_insert(
            csv_path, excel_path,
            csv_cols=["PartCode"], excel_cols=["PartCode"],
            output_path=output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        self.assertEqual(cell.value, "000461200000102")
        self.assertEqual(cell.number_format, "@")
        # Font family + size are forced; italic decoration survives.
        self.assertEqual(cell.font.name, "Times New Roman")
        self.assertEqual(cell.font.size, 9)
        self.assertTrue(cell.font.italic)
        self.assertEqual(cell.alignment.horizontal, "center")


if __name__ == "__main__":
    unittest.main()
